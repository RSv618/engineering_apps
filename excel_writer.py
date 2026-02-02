from openpyxl.chart import LineChart, Reference
from openpyxl.chart.chartspace import ChartSpace
from functools import lru_cache
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D, XDRPoint2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from typing import Any
import collections
from utils import get_dia_code, resource_path
from rebar_optimizer import find_optimized_cutting_plan
from datetime import timedelta

# 1. Save the original to_tree method to preserve standard functionality
_original_to_tree = ChartSpace.to_tree

# 2. Define the patched method
def _patched_to_tree(self, tagname=None, idx=None, namespace=None):
    """
    Intercept the XML generation to force the chart background to be transparent.
    This modifies the <c:chartSpace> element, which is the root container
    responsible for the white background.
    """
    # Ensure the 'spPr' (Shape Properties) attribute exists
    if self.spPr is None:
        self.spPr = GraphicalProperties()

    # Force 'No Fill' (Transparent Background)
    self.spPr.noFill = True

    # Force 'No Line' (No Border)
    if self.spPr.ln is None:
        self.spPr.ln = LineProperties()
    self.spPr.ln.noFill = True

    # 3. Call the original logic to generate the XML tree
    return _original_to_tree(self, tagname, idx, namespace)

# 4. Apply the patch to the class
ChartSpace.to_tree = _patched_to_tree

def excel_col_width_to_px(width: float | None) -> int:
    """
    Approximates the conversion of an openpyxl column width to pixels.

    Args:
        width: The column width in openpyxl units.

    Returns:
        The approximate width in pixels.
    """
    return int((width + 0.71) * 7) if width is not None else int((8.43 + 0.71) * 7)


def excel_row_height_to_px(height: float | None) -> float:
    """
    Converts an openpyxl row height (in points) to pixels.

    Args:
        height: The row height in points.

    Returns:
        The height in pixels.
    """
    return (height if height is not None else 15) * 96 / 72


def center_img(img: Image, cell_ref: str, ws: Worksheet) -> Image:
    """
    Calculates the anchor position to center an image within a worksheet cell.

    Args:
        img: The openpyxl Image object.
        cell_ref: The cell reference string (e.g., 'A3').
        ws: The openpyxl Worksheet object.

    Returns:
        The Image object with its anchor property set.
    """
    cell = ws[cell_ref]
    row, col = cell.row, cell.column

    # --- cumulative Y offset ---
    row_heights = [
        ws.row_dimensions[r].height or ws.sheet_format.defaultRowHeight
        for r in range(1, row)
    ]
    anchor_y = excel_row_height_to_px(sum(row_heights))

    # --- current row height ---
    row_h = excel_row_height_to_px(ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight)

    # --- cumulative X offset ---
    col_widths = [
        ws.column_dimensions[get_col_letter_cached(c)].width or ws.sheet_format.defaultColWidth
        for c in range(1, col)
    ]
    anchor_x = excel_col_width_to_px(sum(col_widths))

    # --- current column width ---
    col_w = excel_col_width_to_px(ws.column_dimensions[get_col_letter_cached(col)].width or ws.sheet_format.defaultColWidth)

    # --- offsets to center ---
    y_offset = max((row_h - img.height) / 2, 0)
    x_offset = max((col_w - img.width) / 2, 0)

    # --- build anchor ---
    pos = XDRPoint2D(pixels_to_EMU(anchor_x + x_offset), pixels_to_EMU(anchor_y + y_offset))
    size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
    img.anchor = AbsoluteAnchor(pos=pos, ext=size)
    return img

def get_canonical_representation(bar: dict[str, Any]) -> tuple[str, tuple]:
    """
    Creates a unique, normalized key for a bar based on its shape and dimensions
    to group identical or mirrored shapes.
    """
    shape = bar['shape']
    dims = bar['shape_dimensions']

    # Normalize shape names for grouping (e.g., tall/wide rectangles are the same shape)
    if 'rectangular' in shape and shape != 'rectangular (diamond)':
        canonical_shape = 'rectangular'
        key = (
            dims.get('A', 0),
            min(dims.get('B', 0), dims.get('C', 0)),
            max(dims.get('B', 0), dims.get('C', 0))
        )
        return canonical_shape, key
    elif shape == 'rectangular (diamond)':
        key = (dims.get('A', 0), dims.get('B', 0))
        return shape, key
    elif shape == 'U':
        key = (
            min(dims.get('A', 0), dims.get('C', 0)),
            dims.get('B', 0),
            max(dims.get('A', 0), dims.get('C', 0))
        )
        return shape, key
    elif shape == 'L':
        key = (dims.get('A', 0), dims.get('B', 0))
        return shape, key
    elif shape == 'flat':
        key = (dims.get('A', 0), dims.get('B', 0))
        return shape, key
    else:
        key = tuple(dims.get(k, 0) for k in sorted(dims.keys()))
        return shape, key

def process_rebar_input(rebar_config: dict[str, Any] | list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Flattens the input dictionary and groups identical bars by shape, dimensions, AND diameter.
    """
    flat_list = []
    if isinstance(rebar_config, dict):
        for bar_type, data in rebar_config.items():
            if bar_type in ['Top Bar', 'Bottom Bar', 'Perimeter Bar']:
                if 'bar_in_x_direction' in data:
                    flat_list.append({'bar_type': bar_type, **data['bar_in_x_direction']})
                if 'bar_in_y_direction' in data:
                    flat_list.append({'bar_type': bar_type, **data['bar_in_y_direction']})
            elif bar_type == 'Vertical Bar':
                flat_list.append({'bar_type': bar_type, **data})
            elif bar_type == 'Stirrups':
                for stirrup in data:
                    flat_list.append({'bar_type': bar_type, **stirrup})
    elif isinstance(rebar_config, list):
        for rebar_config_instance in rebar_config:
            for bar_type, data in rebar_config_instance.items():
                if bar_type in ['Top Bar', 'Bottom Bar', 'Perimeter Bar']:
                    if 'bar_in_x_direction' in data:
                        flat_list.append({'bar_type': bar_type, **data['bar_in_x_direction']})
                    if 'bar_in_y_direction' in data:
                        flat_list.append({'bar_type': bar_type, **data['bar_in_y_direction']})
                elif bar_type == 'Vertical Bar':
                    flat_list.append({'bar_type': bar_type, **data})
                elif bar_type == 'Stirrups':
                    for stirrup in data:
                        flat_list.append({'bar_type': bar_type, **stirrup})
    else:
        raise TypeError(f'Invalid rebar_config type. Expected dict or list, got {type(rebar_config)}')

    grouped_bars = collections.OrderedDict()
    for bar in flat_list:
        canonical_shape, dim_key = get_canonical_representation(bar)
        group_key = (canonical_shape, bar['diameter'], dim_key)

        if group_key not in grouped_bars:
            grouped_bars[group_key] = {
                'original_shape': bar['shape'],
                'bar_types': {bar['bar_type']},
                'quantity': 0,
                'diameter': bar['diameter'],
                'cut_length': bar['total_cut_length_mm'],
                'shape_dimensions': bar['shape_dimensions']
            }

        grouped_bars[group_key]['quantity'] += bar['quantity']
        grouped_bars[group_key]['bar_types'].add(bar['bar_type'])

    processed_list = []
    for data in grouped_bars.values():
        processed_list.append({
            'shape': data['original_shape'],
            'bar_type': ',\n'.join(sorted(list(data['bar_types']))),
            'quantity': data['quantity'],
            'diameter': data['diameter'],
            'cut_length': data['cut_length'],
            'shape_dimensions': data['shape_dimensions']
        })

    return processed_list

def create_excel_cutting_list(rebar_config: dict[str, Any],
                              cuts_by_diameter: dict,
                              market_lengths: dict[str, list],
                              output_filename: str = 'rebar_cutting_schedule.xlsx'):
    """
    Generates a formatted Excel rebar cutting list from a rebar configuration dictionary.
    """
    processed_data = process_rebar_input(rebar_config)
    purchase_list, cutting_plan = find_optimized_cutting_plan(cuts_by_diameter, market_lengths)

    proceed_cutting_plan = True
    for plan in cutting_plan:
        if 'Error' in plan:
            proceed_cutting_plan = False

    max_legs = 0
    if processed_data:
        max_legs = max(len(bar['shape_dimensions']) for bar in processed_data)

    # This list comprehension is creating the dictionary keys 'A', 'B', etc., which is fine.
    dimension_headers = [chr(ord('A') + i) for i in range(max_legs)]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rebar Cutting Schedule'

    # --- Styles ---
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    thick_black_side = Side(style='thick', color='404040')
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    header_border = Border(left=white_side, right=white_side, top=black_side, bottom=black_side)

    # --- Static and Dynamic Headers ---
    static_headers = ['Illustration', 'Bar Type', 'Diameter', 'Quantity', 'Cut Length']
    all_headers = static_headers + dimension_headers

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(static_headers) + max_legs)
    title_cell = ws['A1']
    title_cell.value = 'Rebar Cutting and Bending Schedule'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    if not proceed_cutting_plan:
        for plan in cutting_plan:
            if 'Error' in plan:
                title_cell.comment = Comment(plan['Error'], '✨rs_uy')
    ws.row_dimensions[1].height = 30

    # --- Headers ---
    for col_num, header_text in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.fill = header_fill
        cell.border = header_border

    # Apply black border to left and right outer edges
    cell = ws.cell(row=2, column=1)
    cell.border = Border(left=black_side, top=black_side, right=white_side, bottom=black_side)
    cell = ws.cell(row=2, column=len(all_headers))
    cell.border = Border(left=white_side, top=black_side, right=black_side, bottom=black_side)

    # --- Column Widths ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.row_dimensions[2].height = 25

    num_static_cols = len(static_headers)
    for i in range(max_legs):
        # Calculate column index: static columns + 1 (for 1-based index) + loop index
        col_idx = num_static_cols + 1 + i
        col_letter = get_col_letter_cached(col_idx)
        ws.column_dimensions[col_letter].width = 12

    shape_to_image_map = {
        'straight': 'straight.png',
        'U': 'u.png',
        'L': 'l.png',
        'rectangular': 'rectangular_outer.png',
        'rectangular (tall)': 'rectangular_tall.png',
        'rectangular (wide)': 'rectangular_wide.png',
        'rectangular (diamond)': 'rectangular_diamond.png',
        'octagonal': 'octagon.png'
    }

    # --- Data Rows ---
    current_row = 3
    for bar in processed_data:
        ws.row_dimensions[current_row].height = 75

        # Use the map to get the correct image filename
        image_filename = shape_to_image_map.get(bar['shape'])
        if image_filename:
            try:
                img_path = f'images/{image_filename}'

                with PILImage.open(img_path) as pil_img:
                    original_width, original_height = pil_img.size

                aspect_ratio = original_height / original_width
                target_width = 90
                target_height = int(target_width * aspect_ratio + 0.5)
                if target_height > 90:
                    target_height = 90
                    target_width = int(90 / aspect_ratio + 0.5)

                img = Image(img_path)
                img.width = target_width
                img.height = target_height

                ws.add_image(center_img(img, f'A{current_row}', ws))

            except FileNotFoundError:
                ws.cell(row=current_row, column=1, value='No Image')
        else:
            ws.cell(row=current_row, column=1, value='No Image')

        try:
            val = bar['diameter']
            diameter_str = f'{val:.1f} mm\n({get_dia_code(val)})'
        except KeyError:
            # Fallback if the diameter code is not found
            diameter_str = f'{bar['diameter']:.1f} mm'

        data_to_write = [
            None,
            bar['bar_type'],
            diameter_str,  # Use the newly formatted string
            bar['quantity'],
            round(bar['cut_length'], 1),
        ]

        for letter in dimension_headers:
            data_to_write.append(bar['shape_dimensions'].get(letter, '-'))

        for col_num, value in enumerate(data_to_write, 1):
            if value is None:
                cell = ws.cell(row=current_row, column=col_num)
            else:
                cell = ws.cell(row=current_row, column=col_num, value=value)

            if col_num == 5:  # Cutlength
                max_length = max(market_lengths[get_dia_code(bar['diameter'])])
                if value > max_length * 1000:
                    proceed_cutting_plan = False
                    cell.font = Font(color='FF0000')
                    cell.comment = Comment(f'Splicing required.\nCutting length exceeds available market length of {max_length:}m.\n'
                                    f'Cannot proceed with purchase order analysis.', '✨rs_uy', height=150, width=200)

            # Alternating BG Color Fill
            if current_row % 2 == 0:
                cell.fill = alter_row_fill
            cell.alignment = cell_alignment

            # Borders
            cell.border = border
            if col_num == 6:  #Shape Dimensions
                cell.border = Border(left=thick_black_side, bottom=black_side, top=black_side, right=black_side)

            # Number Format
            if col_num >= 5:
                cell.number_format = '#,##0" mm"'

        current_row += 1

    if proceed_cutting_plan:
        add_sheet_purchase_plan(wb, purchase_list)
        add_sheet_cutting_plan(wb, cutting_plan)
    wb.save(output_filename)
    print(f'Excel sheet {output_filename} has been created successfully.')

def add_sheet_purchase_plan(wb, purchase_list) -> Workbook:
    ws = wb.create_sheet('Rebar Purchase')

    # --- Styles ---
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    header_border = Border(left=white_side, right=white_side, top=black_side, bottom=black_side)

    # --- Static and Dynamic Headers ---
    headers = purchase_list[0].keys()

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws['A1']
    title_cell.value = 'Purchase Qty by Length & Diameter'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # --- Headers ---
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.fill = header_fill
        cell.border = header_border

    # Apply black border to left and right outer edges
    cell = ws.cell(row=2, column=1)
    cell.border = Border(left=black_side, top=black_side, right=white_side, bottom=black_side)
    cell = ws.cell(row=2, column=len(headers))
    cell.border = Border(left=white_side, top=black_side, right=black_side, bottom=black_side)

    # --- Column Widths ---
    ws.column_dimensions['A'].width = 20
    ws.row_dimensions[2].height = 25

    for item in purchase_list:
        ws.append(list(item.values()))

    for current_row in range(3, 3 + len(purchase_list)):
        ws.row_dimensions[current_row].height = 25

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)

            if cell.value == 0:
                cell.value = ''

            # Alternating BG Color Fill
            if current_row % 2 == 0:
                cell.fill = alter_row_fill
            cell.alignment = cell_alignment

            # Borders
            cell.border = border

            # Number Format
            if col_num >= 2:
                cell.number_format = '#,##0'
    return wb

def add_sheet_cutting_plan(wb, cutting_plan) -> Workbook:
    ws = wb.create_sheet('Cutting Plan')

    # --- Styles ---
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    header_border = Border(left=white_side, right=white_side, top=black_side, bottom=black_side)

    # --- Static and Dynamic Headers ---
    headers = ['Diameter', 'Quantity', 'Length', 'Cuts', 'Detailed Instructions']

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws['A1']
    title_cell.value = 'Cutting Plan'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # --- Headers ---
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.fill = header_fill
        cell.border = header_border

    # Apply black border to left and right outer edges
    cell = ws.cell(row=2, column=1)
    cell.border = Border(left=black_side, top=black_side, right=white_side, bottom=black_side)
    cell = ws.cell(row=2, column=len(headers))
    cell.border = Border(left=white_side, top=black_side, right=black_side, bottom=black_side)

    # --- Column Widths ---
    ws.column_dimensions['E'].width = 70
    ws.column_dimensions['D'].width = 15
    ws.row_dimensions[2].height = 25
    col_map = {i:header for i, header in enumerate(headers, 1)}
    for current_row in range(3, 3 + len(cutting_plan)):
        ws.row_dimensions[current_row].height = 50

        for col_num in range(1, len(headers) + 1):
            data = cutting_plan[current_row - 3]
            if col_num < 3:
                cell = ws.cell(row=current_row, column=col_num, value=data[col_map[col_num]])
            elif col_map[col_num] == 'Length':
                cell = ws.cell(row=current_row, column=col_num, value=f'{data[col_map[col_num]]}m')
            elif col_map[col_num] == 'Cuts':
                # get cuts
                cuts = '\n'.join(data['Cut Per RSB'])
                cell = ws.cell(row=current_row, column=col_num, value=cuts)
            else: # col_num == 'Detailed Instructions':
                # get detailed instructions
                qty = data['Quantity']
                length = data['Length']
                dia = data['Diameter']
                cuts = [cut.replace('x', '×') for cut in data['Cut Per RSB']]
                # Cut each of the 4 pcs of 13.5m Ø10 bars into 4×2.095m and 3×1.695m lengths.
                if len(cuts) > 2:
                    cuts = cuts[:-1] + ['and ' + cuts[-1]]
                    cuts = ', '.join(cuts)
                elif len(cuts) == 2:
                    cuts = ' and '.join(cuts)
                else:
                    cuts = cuts[0]

                if qty > 1:
                    value = f'  Cut each of the {qty}pcs of {length}m RSB ({dia}) into {cuts} lengths.'
                else:
                    value = f'  Cut 1pc of {length}m RSB ({dia}) into {cuts} lengths.'
                cell = ws.cell(row=current_row, column=col_num, value=value)

            # Alternating BG Color Fill
            if current_row % 2 == 0:
                cell.fill = alter_row_fill
            if col_num == 5:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = cell_alignment

            # Borders
            cell.border = border
    return wb

def add_sheet_cutting_list(title: str, rebar_config: list[dict[str, Any]],
                              market_lengths: dict[str, list], wb: Workbook) -> tuple[Workbook, bool]:
    """
    Generates a formatted Excel rebar cutting list from a rebar configuration dictionary.
    """
    ws = wb.create_sheet(f'{title} Cutting List')
    proceed_purchase_plan = True

    max_legs = 0
    if rebar_config:
        max_legs = max(len(bar['shape_dimensions']) for bar in rebar_config)

    # This list comprehension is creating the dictionary keys 'A', 'B', etc., which is fine.
    dimension_headers = [chr(ord('A') + i) for i in range(max_legs)]

    # --- Styles ---
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    thick_black_side = Side(style='thick', color='404040')
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    header_border = Border(left=white_side, right=white_side, top=black_side, bottom=black_side)

    # --- Static and Dynamic Headers ---
    static_headers = ['Illustration', 'Bar Type', 'Diameter', 'Quantity', 'Cut Length']
    all_headers = static_headers + dimension_headers

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(static_headers) + max_legs)
    title_cell = ws['A1']
    title_cell.value = f'{title} Rebar Cutting and Bending Schedule'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # --- Headers ---
    for col_num, header_text in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.fill = header_fill
        cell.border = header_border

    # Apply black border to left and right outer edges
    cell = ws.cell(row=2, column=1)
    cell.border = Border(left=black_side, top=black_side, right=white_side, bottom=black_side)
    cell = ws.cell(row=2, column=len(all_headers))
    cell.border = Border(left=white_side, top=black_side, right=black_side, bottom=black_side)

    # --- Column Widths ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.row_dimensions[2].height = 25

    num_static_cols = len(static_headers)
    for i in range(max_legs):
        # Calculate column index: static columns + 1 (for 1-based index) + loop index
        col_idx = num_static_cols + 1 + i
        col_letter = get_col_letter_cached(col_idx)
        ws.column_dimensions[col_letter].width = 12

    shape_to_image_map = {
        'straight': 'straight.png',
        'U': 'u.png',
        'L': 'l.png',
        'rectangular': 'rectangular_outer.png',
        'rectangular (tall)': 'rectangular_tall.png',
        'rectangular (wide)': 'rectangular_wide.png',
        'rectangular (diamond)': 'rectangular_diamond.png',
        'octagonal': 'octagon.png',
        'flat (tall)': 'flat_tall.png',
        'flat (wide)': 'flat_wide.png',
    }

    # --- Data Rows ---
    current_row = 3
    for bar in rebar_config:
        ws.row_dimensions[current_row].height = 75

        # Use the map to get the correct image filename
        image_filename = shape_to_image_map.get(bar['shape'])
        if image_filename:
            try:
                img_path = resource_path(f'images/{image_filename}')

                with PILImage.open(img_path) as pil_img:
                    original_width, original_height = pil_img.size

                aspect_ratio = original_height / original_width
                target_width = 90
                target_height = int(target_width * aspect_ratio + 0.5)
                if target_height > 90:
                    target_height = 90
                    target_width = int(90 / aspect_ratio + 0.5)

                img = Image(img_path)
                img.width = target_width
                img.height = target_height

                ws.add_image(center_img(img, f'A{current_row}', ws))

            except FileNotFoundError:
                ws.cell(row=current_row, column=1, value='No Image')
        else:
            ws.cell(row=current_row, column=1, value='No Image')

        try:
            val = bar['diameter']
            diameter_str = f'{val:.1f} mm\n({get_dia_code(val)})'
        except KeyError:
            # Fallback if the diameter code is not found
            diameter_str = f'{bar['diameter']:.1f} mm'

        data_to_write = [
            None,
            bar['bar_type'],
            diameter_str,  # Use the newly formatted string
            bar['quantity'],
            round(bar['cut_length'], 1),
        ]

        for letter in dimension_headers:
            data_to_write.append(bar['shape_dimensions'].get(letter, '-'))

        for col_num, value in enumerate(data_to_write, 1):
            if value is None:
                cell = ws.cell(row=current_row, column=col_num)
            else:
                cell = ws.cell(row=current_row, column=col_num, value=value)

            if col_num == 5:  # Cutlength
                dia_code = get_dia_code(bar['diameter'])
                # Check if market lengths are defined for this diameter
                if dia_code in market_lengths and market_lengths[dia_code]:
                    max_length = max(market_lengths[dia_code])
                    if value > max_length * 1000:
                        proceed_purchase_plan = False
                        cell.font = Font(color='FF0000')
                        cell.comment = Comment(
                            f'Splicing required.\nCutting length exceeds available market length of {max_length:}m.',
                            '✨rs_uy', height=150, width=200)
                else:
                    # This case handles when validation is bypassed or no lengths are available.
                    proceed_purchase_plan = False
                    cell.font = Font(color='FF0000')
                    cell.comment = Comment(
                        f'No market length selected for this diameter ({dia_code}).\nCannot proceed with purchase plan analysis.',
                        '✨rs_uy', height=150, width=200)

            # Alternating BG Color Fill
            if current_row % 2 == 0:
                cell.fill = alter_row_fill
            cell.alignment = cell_alignment

            # Borders
            cell.border = border
            if col_num == 6:  #Shape Dimensions
                cell.border = Border(left=thick_black_side, bottom=black_side, top=black_side, right=black_side)

            # Number Format
            if col_num >= 5:
                cell.number_format = '#,##0" mm"'

        current_row += 1

    return wb, proceed_purchase_plan

def add_concrete_plan_to_workbook(wb: Workbook, breakdown: list):
    """
    Adds a sheet with a concrete volume breakdown and an EDITABLE mix design section.
    Calculations for materials are done via Excel Formulas so the user can tweak the mix.
    """
    ws = wb.create_sheet('Concrete Purchase')

    # --- Styles ---
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    header_border = Border(left=white_side, right=white_side, top=black_side, bottom=black_side)

    # Style for input cells to indicate they are editable (Light Yellow)
    input_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # =========================================================
    # SECTION 1: VOLUME BREAKDOWN
    # =========================================================
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = 'Concrete Volume Breakdown'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = 'Foundation Type'
    ws['B2'] = 'Vol. Per Footing'
    ws['C2'] = 'Quantity'
    ws['D2'] = 'Vol. Per Type'
    ws.row_dimensions[2].height = 25
    for col in ['A', 'B', 'C', 'D']:
        c = ws[f'{col}2']
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = header_border

    # Apply black border to left and right outer edges
    cell = ws.cell(row=2, column=1)
    cell.border = Border(left=black_side, top=black_side, right=white_side, bottom=black_side)
    cell = ws.cell(row=2, column=4)
    cell.border = Border(left=white_side, top=black_side, right=black_side, bottom=black_side)

    row_idx = 3
    for name, vol, n_footing in breakdown:
        vol_per_footing = vol / n_footing
        ws[f'A{row_idx}'] = name
        ws[f'B{row_idx}'] = vol_per_footing
        ws[f'C{row_idx}'] = n_footing
        ws[f'D{row_idx}'] = f'=B{row_idx}*C{row_idx}'

        ws[f'B{row_idx}'].number_format = f'#,##0.00" m³"'
        ws[f'D{row_idx}'].number_format = f'#,##0.00" m³"'
        ws[f'A{row_idx}'].border = thin_border
        ws[f'B{row_idx}'].border = thin_border
        ws[f'C{row_idx}'].border = thin_border
        ws[f'D{row_idx}'].border = thin_border
        if row_idx %2==0:
            ws[f'A{row_idx}'].fill = alter_row_fill
            ws[f'B{row_idx}'].fill = alter_row_fill
            ws[f'C{row_idx}'].fill = alter_row_fill
            ws[f'D{row_idx}'].fill = alter_row_fill
        row_idx += 1

    # Total Row
    total_vol_row = row_idx
    ws.merge_cells(f'A{row_idx}:C{row_idx}')
    ws[f'A{total_vol_row}'] = 'Total Volume'
    ws[f'D{total_vol_row}'].value = f'=sum(D3:D{total_vol_row-1})'
    ws[f'A{total_vol_row}'].font = Font(bold=True)
    ws[f'D{total_vol_row}'].font = Font(bold=True)
    ws[f'A{total_vol_row}'].border = thin_border
    ws[f'B{total_vol_row}'].border = thin_border
    ws[f'C{total_vol_row}'].border = thin_border
    ws[f'D{total_vol_row}'].border = thin_border
    ws[f'A{total_vol_row}'].alignment = Alignment(horizontal='right')
    ws[f'D{total_vol_row}'].number_format = f'#,##0.00" m³"'

    # Store reference to total volume cell (e.g., 'B10')
    ref_total_vol = f'D{total_vol_row}'

    # =========================================================
    # SECTION 2: EDITABLE MIX PARAMETERS
    # =========================================================
    param_start_row = total_vol_row + 2

    ws.merge_cells(f'A{param_start_row}:B{param_start_row}')
    cell = ws[f'A{param_start_row}']
    cell.value = 'Mix Design Parameters'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[param_start_row].height = 30

    # Parameter Labels and Default Values
    # (Label, Default Value, Key for reference)
    params = [
        ('Cement Ratio', 1, 'ratio_c'),
        ('Sand Ratio', 2, 'ratio_s'),
        ('Gravel Ratio', 3, 'ratio_g'),
        ('Water-Cement Ratio', 0.50, 'wc'),
        ('Bag Weight', 40, 'bag_wt'),
        ('Cement Density', 1440, 'cem_dens'),  # Usually hidden or static, but making it explicit is safer
        ('Dry Volume Factor', 1.54, 'factor'),
        ('Wastage Multiplier', 1.05, 'waste'),
    ]

    # Dictionary to store cell addresses for formulas
    refs = {}

    current_row = param_start_row + 1
    for label, val, key in params:
        # Label
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].border = thin_border

        # Value (Input)
        c = ws[f'B{current_row}']
        c.value = val
        c.alignment = center_align
        c.border = thin_border
        c.fill = input_fill  # Highlight as editable
        if label == 'Bag Weight':
            c.number_format = '#,##0" kg"'
        elif label == 'Cement Density':
            c.number_format = f'#,##0" kg/m³"'
        # Save address (e.g., 'B15')
        refs[key] = f'B{current_row}'

        current_row += 1

    # =========================================================
    # SECTION 3: MATERIAL ESTIMATION (FORMULAS)
    # =========================================================
    est_row = current_row + 1
    ws.merge_cells(f'A{est_row}:B{est_row}')
    cell = ws[f'A{est_row}']
    cell.value = 'Materials for Purchase'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[est_row].height = 30

    headers = ['Material', 'Quantity']
    ws.row_dimensions[est_row + 1].height = 25
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=est_row + 1, column=i)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = header_border

    # Apply black border to left and right outer edges
    cell = ws.cell(row=est_row + 1, column=1)
    cell.border = Border(left=black_side, top=black_side, right=white_side, bottom=black_side)
    cell = ws.cell(row=est_row + 1, column=2)
    cell.border = Border(left=white_side, top=black_side, right=black_side, bottom=black_side)

    # --- FORMULA CONSTRUCTION ---
    # 1. Total Ratio Sum = (Rc + Rs + Rg)
    sum_ratio = f'({refs['ratio_c']} + {refs['ratio_s']} + {refs['ratio_g']})'

    # 2. Total Dry Volume = TotalWet * Factor
    dry_vol = f'({ref_total_vol} * {refs['factor']})'

    # 3. Cement Calculation
    # Vol_Cement_Part = DryVol * (Rc / Sum)
    # Weight_Cement = Vol_Cement_Part * Density
    # Bags = Weight / BagWeight
    # Total = Bags * Wastage
    formula_cement = f'=(({dry_vol} * ({refs['ratio_c']} / {sum_ratio})) * {refs['cem_dens']} / {refs['bag_wt']}) * {refs['waste']}'

    # 4. Sand Calculation
    # Vol_Sand = DryVol * (Rs / Sum) * Wastage
    formula_sand = f'=({dry_vol} * ({refs['ratio_s']} / {sum_ratio})) * {refs['waste']}'

    # 5. Gravel Calculation
    # Vol_Gravel = DryVol * (Rg / Sum) * Wastage
    formula_gravel = f'=({dry_vol} * ({refs['ratio_g']} / {sum_ratio})) * {refs['waste']}'

    # 6. Water Calculation
    # Weight of Cement (Total) = (Bags * BagWeight) ... or derived from formula 3 without bag division
    # Let's derive it cleanly: Weight_Cement = Vol_Cement_Part * Density * Wastage
    # Water Weight = Weight_Cement * WC_Ratio
    # Water Vol (L) = Water Weight (1kg = 1L)
    # Formula:
    cement_weight_kg = f'((({dry_vol} * ({refs['ratio_c']} / {sum_ratio})) * {refs['cem_dens']}) * {refs['waste']})'
    formula_water = f'={cement_weight_kg} * {refs['wc']}'

    materials = [
        ('Cement', formula_cement, 'Bags', 'Based on bag weight & ratio'),
        ('Sand', formula_sand, 'm³', 'Loose Volume'),
        ('Gravel', formula_gravel, 'm³', 'Loose Volume'),
        ('Water', formula_water, 'Liters', 'Based on w/c ratio')
    ]

    r = est_row + 2
    for mat, formula, unit, note in materials:
        # Name
        ws.cell(row=r, column=1, value=mat).border = thin_border

        # Formula Cell
        c = ws.cell(row=r, column=2)
        c.value = formula  # Write the Excel formula string
        c.number_format = f'#,##0.00" {unit}"'
        c.alignment = center_align
        c.border = thin_border

        # Note
        # ws.cell(row=r, column=3, value=note).border = thin_border

        if r%2==0:
            for i in range(len(headers)):
                c = ws.cell(row=r, column=i + 1)
                c.fill = alter_row_fill
        r += 1

    # Add Footer Note
    cell = ws.cell(row=r, column=1, value='Note: Modify values in yellow cells to update quantities automatically.')
    cell.font = Font(name='Calibri', size=10, color='5D5D5D')

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def delete_blank_worksheets(wb: Workbook) -> Workbook:
    """
    Deletes all blank worksheets from an Excel workbook.

    Args:
        wb: the Excel workbook.
    """
    sheets_to_delete = []

    for sheet in wb.worksheets:
        # Check if the sheet contains any data
        # max_row and max_column will be 1 if the sheet is empty
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            sheets_to_delete.append(sheet)

    for sheet in sheets_to_delete:
        wb.remove(sheet)

    return wb

@lru_cache(maxsize=256)
def get_col_letter_cached(col_idx):
    return get_column_letter(col_idx)

def get_range(row_from=1, row_to=1, col_from=1, col_to=1):
    return f'{get_col_letter_cached(col_from)}{row_from}:{get_col_letter_cached(col_to)}{row_to}'

def set_style_to_range(ws, cell_range, font:Font=None, fill: PatternFill=None, border: Border=None, alignment: Alignment=None,
                       number_format: str = None):
    for row in ws[cell_range]:
        for cell in row:
            if font is not None:
                cell.font = font
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment
            if number_format is not None:
                cell.number_format = number_format

def create_schedule_sheet(ws, data, check_boxes, table_start_date, total_days):
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    dash_side = Side(style='dashed', color='888888')
    none_side = Side(style='none')
    small_font = Font(name='Calibri', size=8)
    smallest_font = Font(name='Calibri', size=5)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    subheader_font = Font(name='Calibri', size=8, bold=False, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    all_white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
    all_black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    top_black_border = Border(top=black_side, left=white_side, right=white_side, bottom=white_side)
    bottom_black_border = Border(bottom=black_side, left=white_side, right=white_side, top=white_side)
    right_black_border = Border(right=black_side, left=white_side, bottom=white_side, top=white_side)
    left_black_border = Border(left=black_side, right=white_side, bottom=white_side, top=white_side)
    ur_black_border = Border(right=black_side, top=black_side, bottom=white_side, left=white_side)
    ul_black_border = Border(left=black_side, top=black_side, bottom=white_side, right=white_side)
    lr_black_border = Border(right=black_side, bottom=black_side, top=white_side, left=white_side)
    ll_black_border = Border(left=black_side, bottom=black_side, top=white_side, right=white_side)
    all_none_border = Border(left=none_side, right=none_side, top=none_side, bottom=none_side)
    top_black_none_border = Border(top=black_side, left=none_side, right=none_side, bottom=none_side)
    bottom_black_none_border = Border(bottom=black_side, left=none_side, right=none_side, top=none_side)
    right_black_none_border = Border(right=black_side, left=none_side, bottom=none_side, top=none_side)
    left_black_none_border = Border(left=black_side, right=none_side, bottom=none_side, top=none_side)
    ur_black_none_border = Border(right=black_side, top=black_side, bottom=none_side, left=none_side)
    ul_black_none_border = Border(left=black_side, top=black_side, bottom=none_side, right=none_side)
    lr_black_none_border = Border(right=black_side, bottom=black_side, top=none_side, left=none_side)
    ll_black_none_border = Border(left=black_side, bottom=black_side, top=none_side, right=none_side)
    ul_dash_border = Border(left=dash_side, top=black_side, bottom=none_side, right=none_side)
    left_dash_border = Border(left=dash_side, top=none_side, bottom=none_side, right=none_side)
    ll_dash_border = Border(left=dash_side, bottom=black_side, top=none_side, right=none_side)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    bar_colors = {
        'Original': '009580',
        'Schedule': '009580',
        'Revised': 'FFC600',
        'Actual': 'FF003C'
    }
    bar_fill_color = {}
    for color in bar_colors:
        bar_fill_color[color] = PatternFill(start_color=bar_colors[color], end_color=bar_colors[color], fill_type='solid')

    headers = ['Activity', 'Weight', 'Schedule']
    if not check_boxes['S-Curve']:
        headers.remove('Weight')
    col_widths = {'Activity': 40, 'Weight': 10, 'Schedule': 10, 'Revised': 10, 'Actual': 10, 'Original': 10,
                  'Start': 10, 'End': 10, 'Days': 5}
    rows_per_activity = 3
    if check_boxes['Revised']:
        headers.append('Revised')
        rows_per_activity += 1
    if check_boxes['Actual']:
        headers.append('Actual')
        rows_per_activity += 1
    if check_boxes['Actual'] or check_boxes['Revised']:
        headers[headers.index('Schedule')] = 'Original'

    # Define table upper left border
    _col_left_table = 2
    _row_top_table = 2
    if _col_left_table > 1:
        for i in range(1, _col_left_table):
            ws.column_dimensions[get_col_letter_cached(i)].width = 1
    if _row_top_table > 1:
        for i in range(1, _row_top_table):
            ws.row_dimensions[i].height = 5

    # Left headers
    i = 0
    for header in headers:
        if header not in ['Original', 'Revised', 'Actual', 'Schedule']:
            ws.merge_cells(start_row=_row_top_table, start_column=_col_left_table + i,
                           end_row=_row_top_table + 2, end_column=_col_left_table + i)
            cell = ws.cell(row=_row_top_table, column=_col_left_table + i, value=header)
            i += 1
        else:
            ws.merge_cells(start_row=_row_top_table, start_column=_col_left_table + i,
                           end_row=_row_top_table, end_column=_col_left_table + i + 2)
            ws.merge_cells(start_row=_row_top_table + 1, start_column=_col_left_table + i,
                           end_row=_row_top_table + 2, end_column=_col_left_table + i)
            ws.merge_cells(start_row=_row_top_table + 1, start_column=_col_left_table + i + 1,
                           end_row=_row_top_table + 2, end_column=_col_left_table + i + 1)
            ws.merge_cells(start_row=_row_top_table + 1, start_column=_col_left_table + i + 2,
                           end_row=_row_top_table + 2, end_column=_col_left_table + i + 2)
            cell = ws.cell(row=_row_top_table, column=_col_left_table + i, value=header)
            ws.cell(row=_row_top_table + 1, column=_col_left_table + i, value='Start')
            ws.cell(row=_row_top_table + 1, column=_col_left_table + i + 1, value='End')
            ws.cell(row=_row_top_table + 1, column=_col_left_table + i + 2, value='Days')
            ws.column_dimensions[get_col_letter_cached(cell.column)].width = col_widths['Start']
            ws.column_dimensions[get_col_letter_cached(cell.column + 1)].width = col_widths['End']
            ws.column_dimensions[get_col_letter_cached(cell.column + 2)].width = col_widths['Days']
            i += 3
        ws.column_dimensions[get_col_letter_cached(cell.column)].width = col_widths[header]

    # Date headers
    # col_width = max(100/total_days,0.12)
    col_width = 0.12  # fixed so that the chart width can easily be estimated
    month_start = []
    year_start = []
    _row_bot_header = _row_top_table + 2
    _col_left_chart = _col_left_table + i
    _col_right_table = _col_left_chart + total_days - 1
    for j in range(total_days):
        date_j = table_start_date + j * timedelta(days=1)
        ws.cell(row=_row_bot_header, column=_col_left_chart + j, value=date_j)
        if j == 0:
            year_start.append(j)
            month_start.append(j)
        elif (date_j - timedelta(days=1)).day > date_j.day:
            month_start.append(j)
            if (date_j - timedelta(days=1)).year < date_j.year:
                year_start.append(j)
        elif j == total_days - 1:
            year_start.append(j + 1)
            month_start.append(j + 1)
        ws.column_dimensions[get_col_letter_cached(_col_left_chart +j)].width = col_width

    # Store columns
    column_idx = {header: 0 for header in headers}
    for i in range(_col_left_table, _col_left_chart):
        cell_value = ws.cell(row=_row_top_table, column=i).value
        if cell_value in column_idx.keys():
            column_idx[cell_value] = i
    schedule_cols = {key: value for key, value in column_idx.items() if key in ['Schedule', 'Original', 'Revised', 'Actual']}

    # Insert months
    prev_month_start = 0
    for j in month_start[1:]:
        ws.merge_cells(start_row=_row_top_table+1, start_column=_col_left_chart+prev_month_start,
                       end_row=_row_top_table+1, end_column=_col_left_chart+j - 1)
        date_j = table_start_date + prev_month_start * timedelta(days=1)
        ws.cell(row=_row_top_table+1, column=_col_left_chart+prev_month_start, value=date_j.strftime('%b'))
        prev_month_start = j

    # Insert years
    prev_year_start = 0
    for j in year_start[1:]:
        ws.merge_cells(start_row=_row_top_table, start_column=_col_left_chart + prev_year_start,
                       end_row=_row_top_table, end_column=_col_left_chart + j - 1)
        date_j = table_start_date + prev_year_start * timedelta(days=1)
        ws.cell(row=_row_top_table, column=_col_left_chart + prev_year_start, value=date_j.year)
        prev_year_start = j

    # Style headers
    for i in range(_row_top_table, _row_bot_header + 1):
        for j in range(_col_left_table, _col_right_table + 1):
            cell = ws.cell(row=i, column=j)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

            if i == _row_top_table:  # Apply top border to the top row
                if j == _col_left_table:
                    cell.border = ul_black_border
                elif j == _col_right_table:
                    cell.border = ur_black_border
                else:
                    cell.border = top_black_border
            elif i == _row_bot_header - 1:  # Apply bottom border to the bottom row
                if j >= _col_left_chart:
                    cell.font = subheader_font

                if j == _col_left_table:
                    cell.border = ll_black_border
                elif j == _col_right_table:
                    cell.border = lr_black_border
                else:
                    cell.border = bottom_black_border
            elif j == _col_left_table:
                cell.border = left_black_border
            elif j == _col_right_table:
                cell.border = right_black_border
            else:
                cell.border = all_white_border
    ws.row_dimensions[_row_bot_header].hidden = True

    # Insert Rows
    _row_top_item = _row_bot_header + 1
    _current_row = _row_top_item
    alter_row_fill_trigger = False
    for row in data:
        j = 0
        for header in headers:
            ws.merge_cells(start_row=_current_row, start_column=_col_left_table + j,
                           end_row=_current_row+rows_per_activity-1, end_column=_col_left_table + j)
            # Activity
            if header == 'Activity':
                cell = ws.cell(row=_current_row, column=_col_left_table + j, value=row['name'])

                # style
                cell.alignment = align_left
                if alter_row_fill_trigger:
                    cell.fill = alter_row_fill
                else:
                    cell.fill = white_fill
                for idx in range(rows_per_activity):
                    cell = ws.cell(row=_current_row + idx, column=_col_left_table + j)
                    cell.border = all_black_border

                j += 1
            elif header == 'Weight':
                cell = ws.cell(row=_current_row, column=_col_left_table + j, value=row['weight'])

                # style
                cell.alignment = align_center
                if alter_row_fill_trigger:
                    cell.fill = alter_row_fill
                else:
                    cell.fill = white_fill
                cell.number_format = '#,##0.00'
                cell.font = small_font
                for idx in range(rows_per_activity):
                    cell = ws.cell(row=_current_row + idx, column=_col_left_table + j)
                    cell.border = all_black_border

                j += 1
            elif header in ['Original', 'Schedule', 'Revised', 'Actual']:
                if header in ['Original', 'Schedule']:
                    start_date, end_date = row['orig']
                elif header == 'Revised':
                    start_date, end_date = row['rev']
                else:  # Actual
                    start_date = ''
                    end_date = ''

                end_date_cell = f'{get_col_letter_cached(_col_left_table + j + 1)}{_current_row}'
                start_date_cell = f'{get_col_letter_cached(_col_left_table + j)}{_current_row}'

                start_cell = ws.cell(row=_current_row, column=_col_left_table + j, value=start_date)
                end_cell = ws.cell(row=_current_row, column=_col_left_table + j + 1, value=end_date)
                day_cell = ws.cell(row=_current_row, column=_col_left_table + j + 2,
                                   value=f'=if(and(len({end_date_cell})>0, len({start_date_cell})>0, '
                                         f'ISNUMBER({end_date_cell}), ISNUMBER({start_date_cell})), '
                                         f'{end_date_cell}-{start_date_cell}+1, "")')

                ws.merge_cells(start_row=_current_row, start_column=_col_left_table + j + 1,
                               end_row=_current_row + rows_per_activity - 1, end_column=_col_left_table + j + 1)
                ws.merge_cells(start_row=_current_row, start_column=_col_left_table + j + 2,
                               end_row=_current_row + rows_per_activity - 1, end_column=_col_left_table + j + 2)

                # style
                for cell in [start_cell, end_cell, day_cell]:
                    cell.alignment = align_center
                    if alter_row_fill_trigger:
                        cell.fill = alter_row_fill
                    else:
                        cell.fill = white_fill
                    cell.font = small_font
                start_cell.number_format = 'mmm dd, yyyy'
                end_cell.number_format = 'mmm dd, yyyy'
                day_cell.number_format = '#,##'
                for idx in range(rows_per_activity):
                    start_cell = ws.cell(_current_row + idx, column=_col_left_table + j)
                    start_cell.border = all_black_border
                    end_cell = ws.cell(_current_row + idx, column=_col_left_table + j + 1)
                    end_cell.border = all_black_border
                    day_cell = ws.cell(_current_row + idx, column=_col_left_table + j + 2)
                    day_cell.border = all_black_border

                j += 3
        _current_row += rows_per_activity
        if alter_row_fill_trigger:
            alter_row_fill_trigger = False
        else:
            alter_row_fill_trigger = True
    _row_bot_table = _current_row - 1

    # Row heights for spacer
    i = _row_bot_header + 1
    while i <= _row_bot_table:
        for j in range(rows_per_activity):
            if j == 0 or (j == rows_per_activity - 1):
                ws.row_dimensions[i].height = 5
            i += 1

    # Actual Sheet
    n_progress_entry = 250
    if check_boxes['Actual']:
        actual_ws = create_input_actual_sheet(ws.parent, data, _col_left_table, _row_top_table, rows_per_activity,
                                              bar_colors, col_widths, n_progress_entry)
        a_sheet = f'\'{actual_ws.title}\'!'
        progress_col_left = get_col_letter_cached(column_idx['Activity'] + 1)
        progress_col_right = get_col_letter_cached(column_idx['Activity'] + n_progress_entry)


    # Style and border the bar-chart
    _row_top_chart = _row_bot_header + 1
    i = 0
    schedule_rows = {key: set() for key in schedule_cols.keys()}
    alter_row_fill_trigger = True
    while i + _row_top_chart <= _row_bot_table:
        if i >= rows_per_activity:
            remainder = i % rows_per_activity
        else:
            remainder = i
        if remainder == 0:
            # toggle alter row
            if alter_row_fill_trigger:
                alter_row_fill_trigger = False
            else:
                alter_row_fill_trigger = True

        for j in range(_col_left_chart, _col_right_table + 1):
            cell = ws.cell(row=i + _row_top_chart, column=j)
            cell.number_format = '""'
            if alter_row_fill_trigger:
                cell.fill = alter_row_fill
            else:
                cell.fill = white_fill

            if remainder == 0:  # Apply top border to the top row
                if j == _col_left_chart:
                    cell.border = ul_black_none_border
                elif j == _col_right_table:
                    cell.border = ur_black_none_border
                else:
                    cell.border = top_black_none_border
            elif remainder == rows_per_activity - 1:  # Apply bottom border to the bottom row
                if j == _col_left_chart:
                    cell.border = ll_black_none_border
                elif j == _col_right_table:
                    cell.border = lr_black_none_border
                else:
                    cell.border = bottom_black_none_border
            elif j == _col_left_table:  # Apply left border to the leftmost column
                cell.border = left_black_none_border
            elif j == _col_right_table:  # Apply right border to the rightmost column
                cell.border = right_black_none_border
            else:
                cell.border = all_none_border

            for idx, key in enumerate(schedule_cols.keys()):
                if remainder == idx + 1:  # Apply formula for
                    if key in ['Original', 'Schedule', 'Revised', 'Actual']:
                        schedule_rows[key].add(i + _row_top_chart)
                        current_date = f'{get_col_letter_cached(j)}${_row_bot_header}'
                        section_row = i + _row_top_chart - remainder
                        if key == 'Actual':
                            dates_row = section_row + rows_per_activity - 2
                            progress_dates_range = (f'{a_sheet}${progress_col_left}${dates_row}:'
                                                    f'${progress_col_right}${dates_row}')
                            progress_pcent_range = (f'{a_sheet}${progress_col_left}${section_row}:'
                                                    f'${progress_col_right}${section_row}')
                            first_progress = f'{a_sheet}${progress_col_left}${section_row}'
                            first_date = f'{a_sheet}${progress_col_left}${dates_row}'
                            last_progress = f'INDEX({progress_pcent_range},1,COUNT({progress_dates_range}))'
                            last_date = f'INDEX({progress_dates_range},1,COUNT({progress_dates_range}))'

                            if check_boxes['S-Curve']:
                                before_index = f'MATCH({current_date}-1/24,{progress_dates_range},1)'
                                after_index = f'{before_index}+1'
                                before_progress = f'INDEX({progress_pcent_range},1,{before_index})'
                                before_date = f'INDEX({progress_dates_range},1,{before_index})'
                                after_progress = f'INDEX({progress_pcent_range},1,{after_index})'
                                after_date = f'INDEX({progress_dates_range},1,{after_index})'
                                increment_progress = f'({after_progress}-{before_progress})/({after_date}-{before_date})'
                                weight = f'${get_col_letter_cached(column_idx['Weight'])}{section_row}'
                                total_weight = f'${get_col_letter_cached(column_idx['Weight'])}{_row_bot_table + 1}'
                                ifs = f'=IF(OR(LEN(TRIM({first_date}))=0,{current_date}<{first_date},{current_date}>TODAY()),""'
                                ifs = f'{ifs},IF({current_date}={first_date},{first_progress}*{weight}/{total_weight}'
                                ifs = f'{ifs},IF({current_date}>{last_date},IF({last_progress}>=1,"",0)'
                                ifs = f'{ifs},IFERROR({increment_progress}*{weight}/{total_weight},""))))'
                                cell.value = ifs
                            else:
                                ifs = (f'=IF(OR(LEN(TRIM({first_date}))=0,{current_date}<{first_date},{current_date}>TODAY()),'
                                       f'"",IF({current_date}>{last_date},IF({last_progress}>=1,"",1),1))')
                                cell.value = ifs

                            # Update start and end dates
                            ws.cell(row=section_row, column=column_idx['Actual']).value = \
                                f'=IF(OR(LEN(TRIM({first_date}))=0, NOT(ISNUMBER({first_date}))), "", {first_date})'
                            cell_start = f'{get_col_letter_cached(column_idx['Actual'])}{section_row}'
                            ws.cell(row=section_row, column=column_idx['Actual'] + 1).value = \
                                f'=IF(LEN({cell_start})=0, "", IF({last_progress}>=1, {last_date}, "On-going"))'
                        else:
                            date_start = f'${get_col_letter_cached(schedule_cols[key])}{section_row}'
                            date_end = f'${get_col_letter_cached(schedule_cols[key] + 1)}{section_row}'
                            duration = f'${get_col_letter_cached(schedule_cols[key] + 2)}{section_row}'
                            if check_boxes['S-Curve']:
                                weight = f'${get_col_letter_cached(column_idx['Weight'])}{section_row}'
                                total_weight = f'${get_col_letter_cached(column_idx['Weight'])}{_row_bot_table + 1}'
                                weighted_value = f'{weight}/({total_weight}*{duration})'
                            else:
                                weighted_value = '1'
                            cell.value = (f'=iferror(if(and({current_date}>={date_start}, {current_date}<={date_end}), '
                                          f'{weighted_value}, ""), "")')
        i += 1

    # Style bar-chart via conditional formatting
    for key in schedule_cols.keys():
        multi_cell_range = []
        first_cell = None
        for row in sorted(list(schedule_rows[key])):
            if first_cell is None:
                first_cell = f'{get_col_letter_cached(_col_left_chart)}{row}'
            multi_cell_range.append(f'${get_col_letter_cached(_col_left_chart)}${row}:${get_col_letter_cached(_col_right_table)}${row}')
        rule = FormulaRule(formula=[f'LEN(TRIM({first_cell}))>0'], stopIfTrue=True, fill=bar_fill_color[key])
        ws.conditional_formatting.add(' '.join(multi_cell_range), rule)

        # Border for each month
        section_rows = schedule_rows[list(schedule_rows.keys())[0]]
        section_rows = [row - 1 for row in section_rows]
        for i in range(_row_top_chart, _row_bot_table + 1):
            for j in month_start[1:-1]:
                cell = ws.cell(row=i, column=j + _col_left_chart)
                if i in section_rows:
                    cell.border = ul_dash_border
                elif (i + 1 in section_rows) or (i + 1 == _row_bot_table + 1):
                    cell.border = ll_dash_border
                else:
                    cell.border = left_dash_border

        # Color the header for legend
        for schedule in schedule_cols.keys():
            cell = ws.cell(row=_row_top_table, column=column_idx[schedule])
            cell.fill = bar_fill_color[schedule]
            if schedule in ['Schedule', 'Original', 'Revised', 'Actual']:  # Hide Days column
                ws.column_dimensions[get_col_letter_cached(column_idx[schedule] + 2)].hidden = True

    # No footer if no S-Curve
    if not check_boxes['S-Curve']:
        return ws

    # Footer
    ws.column_dimensions[get_col_letter_cached(column_idx['Weight'])].hidden = True
    _row_top_footer = _row_bot_table + 1
    ws.cell(row=_row_top_footer, column=column_idx['Weight'],
            value=f'=sum({get_col_letter_cached(column_idx['Weight'])}'
                  f'{_row_bot_header+1}:{get_col_letter_cached(column_idx['Weight'])}{_row_bot_table})')
    for idx, (schedule, rows) in enumerate(schedule_rows.items()):
        ws.merge_cells(start_row=_row_top_footer + idx, end_row=_row_top_footer + idx,
                       start_column=column_idx['Weight'] + 1, end_column=_col_left_chart - 1)
        ws.merge_cells(start_row=_row_top_footer + idx + len(schedule_rows), end_row=_row_top_footer + idx + len(schedule_rows),
                       start_column=column_idx['Weight'] + 1, end_column=_col_left_chart - 1)
        if len(schedule_rows) > 1:
            label = f'Monthly Progress [{schedule}] →'
        else:
            label = f'Monthly Progress →'
        cell = ws.cell(row=_row_top_footer + idx, column=column_idx['Weight'] + 1, value=label)
        cell.font = small_font
        cell.alignment = align_right
        cell = ws.cell(row=_row_top_footer + idx + len(schedule_rows), column=column_idx['Weight'] + 1, value=f'{schedule} (Daily)')
        cell.alignment = align_right
    for i in range(_col_left_chart, _col_right_table + 1):
        for idx, (schedule, rows) in enumerate(schedule_rows.items()):
            idx += len(schedule_rows)
            col = get_col_letter_cached(i)
            cell_ranges = [f'{col}{row}' for row in rows]
            if i == _col_left_chart:
                ws.cell(row=_row_top_footer + idx, column=i, value=f'=sum({','.join(cell_ranges)})')
            else:
                prev_cell = f'{get_col_letter_cached(i - 1)}{_row_top_footer + idx}'
                ws.cell(row=_row_top_footer + idx, column=i, value=f'=sum({prev_cell}, {','.join(cell_ranges)})')
    prev_month = 0
    for i in month_start[1:]:
        for idx in range(len(schedule_rows)):
            ws.merge_cells(start_row=_row_top_footer + idx, end_row=_row_top_footer + idx,
                           start_column=_col_left_chart+prev_month, end_column=_col_left_chart+i - 1)
            ws.cell(row=_row_top_footer + idx, column=_col_left_chart + prev_month,
                    value=f'={get_col_letter_cached(_col_left_chart+i-1)}{_row_top_footer+idx+len(schedule_rows)}')
        prev_month = i

    # Style footer
    _row_bot_footer = _row_top_footer + 2*len(schedule_rows) - 1
    ws.merge_cells(start_row=_row_top_footer, end_row=_row_bot_footer,
                   start_column=column_idx['Activity'], end_column=column_idx['Activity'])
    ws.merge_cells(start_row=_row_top_footer, end_row=_row_bot_footer,
                   start_column=column_idx['Weight'], end_column=column_idx['Weight'])
    cell_range = get_range(row_from=_row_top_footer, row_to=_row_bot_footer, col_from=_col_left_table, col_to=_col_right_table)
    set_style_to_range(ws, cell_range, border=all_black_border, alignment=align_center, fill=grey_fill)
    cell_range = get_range(row_from=_row_bot_footer + 1, row_to=_row_bot_footer + 1, col_from=_col_left_table, col_to=_col_right_table)
    set_style_to_range(ws, cell_range, border=top_black_none_border)
    cell_range = get_range(row_from=_row_top_footer, row_to=_row_bot_footer, col_from=column_idx['Weight'], col_to=column_idx['Weight'])
    set_style_to_range(ws, cell_range, number_format='#,##0.00', font=small_font)
    cell_range = get_range(row_from=_row_top_footer, row_to=_row_bot_footer, col_from=column_idx['Weight'] + 1, col_to=_col_right_table)
    set_style_to_range(ws, cell_range, font=small_font)
    cell_range = get_range(row_from=_row_top_footer, row_to=_row_bot_footer, col_from=column_idx['Weight'] + 1, col_to=column_idx['Weight'] + 1)
    set_style_to_range(ws, cell_range, alignment=align_right)
    cell_range = get_range(row_from=_row_top_footer, row_to=_row_bot_footer, col_from=_col_left_chart, col_to=_col_right_table)
    if col_width <= 0.15:
        set_style_to_range(ws, cell_range, number_format='0%', font=smallest_font)
    else:
        set_style_to_range(ws, cell_range, number_format='0%')

    # Hide daily rows
    for i in range(len(schedule_rows)):
        ws.row_dimensions[_row_bot_footer-i].height = 1

    # S-Curve
    chart = LineChart()

    # Remove default style (this can override fill settings)
    chart.style = None

    # Chart colors
    chart_line_colors = ['00b49c', 'e7b200', 'ff6a73']

    # Add data
    chart_data = Reference(ws,
                     min_col=column_idx['Weight'] + 1,
                     max_col=_col_right_table,
                     min_row=_row_top_footer + len(schedule_rows),
                     max_row=_row_top_footer + 2 * len(schedule_rows) - 1)

    chart.add_data(chart_data, titles_from_data=True, from_rows=True)

    # Apply colors to each series
    for idx, series in enumerate(chart.series):
        if idx < len(chart_line_colors):
            series.graphicalProperties.line.solidFill = chart_line_colors[idx]
            series.graphicalProperties.line.width = 25000
            series.marker.symbol = None

    # 1. Remove Inner Plot Area Background (The area behind the lines)
    # We still need this because the patch only handles the Outer Container.
    chart.plot_area.graphicalProperties = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))

    # 2. Remove Gridlines, Axes, Legend, Title
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.y_axis.delete = True
    chart.x_axis.delete = True
    chart.legend = None
    chart.title = None

    # 3. Rounded Corners
    chart.roundedCorners = False

    # --- POSITIONING ---
    # 1. Define Chart Dimensions (in EMUs)
    # We must explicitly set the size because we are using a custom anchor.
    # 1 cm = 360,000 EMUs. Adjust these CM values to match your desired size.
    # (Your previous logic used approx width=20, height=10. 20cm is roughly 7.8 inches)
    num_cols = (_col_right_table-_col_left_chart+1)
    chart_col_width = col_width * num_cols
    width_emu = cm_to_EMU(0.1758*chart_col_width+0.9438)

    height_per_section = rows_per_activity * ws.sheet_format.defaultRowHeight + 5 * 2
    num_section = len(data)
    if height_per_section == 85:
        emu_height = 0.0219*(height_per_section*num_section)+0.7831
    elif height_per_section == 70:
        emu_height = 0.0193*(height_per_section*num_section)+0.8076
    elif height_per_section == 55:
        emu_height = 0.0154*(height_per_section*num_section)+0.7809
    else:
        emu_height = 0.02*(height_per_section*num_section)-0.0854
    height_emu = cm_to_EMU(emu_height)

    # Create the size object required by OneCellAnchor
    size = XDRPositiveSize2D(cx=width_emu, cy=height_emu)

    # 2. Define the Anchor Marker
    # We anchor to the cell DIAGONALLY ABOVE-LEFT of where the chart starts.
    # This allows us to use positive offsets to shift the chart 'Left' or 'Up'
    # relative to the target start cell.

    # _col_left_chart is 1-based.
    # Subtract 1 to get 0-based index. Subtract 1 more to get the previous column.
    anchor_col = _col_left_chart
    anchor_row = _row_top_chart

    # 3. Define Offsets (The 'Shift')
    # These offsets are from the top-left of the ANCHOR cell (the previous cell).
    # To shift the chart Left: reduce offset_x.
    # To shift the chart Right: increase offset_x.
    # To shift the chart Up: reduce offset_y.
    offset_x = pixels_to_EMU(-19)  # Move 40px right from the previous column (Pushing it into place)
    offset_y = pixels_to_EMU(-20)  # Move 10px down from the previous row

    marker = AnchorMarker(col=anchor_col, colOff=offset_x, row=anchor_row, rowOff=offset_y)

    # 4. Apply Anchor and Add
    chart.anchor = OneCellAnchor(_from=marker, ext=size)

    # IMPORTANT: Do not pass a cell address (e.g. 'A1') as the second argument.
    # Passing only the chart forces openpyxl to use the 'chart.anchor' we just created.
    ws.add_chart(chart)

    return ws

def create_input_actual_sheet(wb, data, _col_left_table, _row_top_table, rows_per_activity, bar_colors, col_widths, n_progress_entry):
    ws = wb.create_sheet('Input Actual')
    ws.sheet_properties.tabColor = bar_colors['Actual']

    # styles
    white_side = Side(style='thin', color='FFFFFF')
    black_side = Side(style='thin', color='404040')
    none_side = Side(style='none')
    small_font = Font(name='Calibri', size=8)
    percent_font = Font(name='Calibri', size=10, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    alter_row_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    all_white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
    all_black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    no_top_black_border = Border(left=black_side, right=black_side, top=none_side, bottom=black_side)
    no_bottom_black_border = Border(left=black_side, right=black_side, top=black_side, bottom=none_side)
    top_black_border = Border(top=black_side, left=white_side, right=white_side, bottom=white_side)
    bottom_black_border = Border(bottom=black_side, left=white_side, right=white_side, top=white_side)
    right_black_border = Border(right=black_side, left=white_side, bottom=white_side, top=white_side)
    left_black_border = Border(left=black_side, right=white_side, bottom=white_side, top=white_side)
    ur_black_border = Border(right=black_side, top=black_side, bottom=white_side, left=white_side)
    ul_black_border = Border(left=black_side, top=black_side, bottom=white_side, right=white_side)
    lr_black_border = Border(right=black_side, bottom=black_side, top=white_side, left=white_side)
    ll_black_border = Border(left=black_side, bottom=black_side, top=white_side, right=white_side)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top = Alignment(horizontal='center', vertical='top')
    align_bottom = Alignment(horizontal='center', vertical='bottom')

    # Define table coner
    if _col_left_table > 1:
        for i in range(1, _col_left_table):
            ws.column_dimensions[get_col_letter_cached(i)].width = 1
    if _row_top_table > 1:
        for i in range(1, _row_top_table):
            ws.row_dimensions[i].height = 5

    # Define headers
    _col_right_table = _col_left_table + n_progress_entry
    for j in range(_col_left_table, _col_right_table + 1):
        ws.merge_cells(start_row=_row_top_table, start_column=j,
                              end_row=_row_top_table + 2, end_column=j)
        if j == _col_left_table:
            value = 'Activity'
            ws.column_dimensions[get_col_letter_cached(j)].width = col_widths['Activity']
        else:
            value = f'Progress\nEntry {j - _col_left_table}'
            ws.column_dimensions[get_col_letter_cached(j)].width = 11
        cell = ws.cell(row=_row_top_table, column=j, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        for i in range(_row_top_table, _row_top_table + 3):
            cell = ws.cell(row=i, column=j)
            cell.border = all_white_border

    # Correct the outer borders
    for j in range(_col_left_table + 1, _col_right_table):
        ws.cell(row=_row_top_table, column=j).border = top_black_border
        ws.cell(row=_row_top_table + 2, column=j).border = bottom_black_border
    ws.cell(row=_row_top_table, column=_col_left_table).border = ul_black_border
    ws.cell(row=_row_top_table + 1, column=_col_left_table).border = left_black_border
    ws.cell(row=_row_top_table + 2, column=_col_left_table).border = ll_black_border
    ws.cell(row=_row_top_table, column=_col_right_table).border = ur_black_border
    ws.cell(row=_row_top_table + 1, column=_col_right_table).border = right_black_border
    ws.cell(row=_row_top_table + 2, column=_col_right_table).border = lr_black_border
    ws.row_dimensions[_row_top_table + 2].hidden = True

    # Add Rows
    i = _row_top_table + 3
    alter_row_fill_trigger = False
    for row in data:
        ws.merge_cells(start_row=i, end_row=i + rows_per_activity -1, start_column=_col_left_table, end_column=_col_left_table)
        cell = ws.cell(row=i, column=_col_left_table, value = row['name'])
        cell.alignment = align_left
        if alter_row_fill_trigger:
            cell.fill = alter_row_fill
            alter_row_fill_trigger = False
        else:
            cell.fill = white_fill
            alter_row_fill_trigger = True
        for idx in range(rows_per_activity):
            cell = ws.cell(row=i, column=_col_left_table)
            cell.border = all_black_border

            if idx == 0 or idx == rows_per_activity - 1:
                ws.row_dimensions[i].height = 5
            i += 1

    # Add entry blocks
    i = _row_top_table + 3
    alter_row_fill_trigger = False
    for _ in range(len(data)):
        for j in range(_col_left_table + 1, _col_right_table + 1):
            date_input_row = i + rows_per_activity - 2
            ws.merge_cells(start_row=i, end_row=date_input_row - 1, start_column=j, end_column=j)
            ws.merge_cells(start_row=date_input_row, end_row=i + rows_per_activity - 1, start_column=j, end_column=j)
            for idx in range(rows_per_activity):
                cell = ws.cell(row=i + idx, column=j)
                if alter_row_fill_trigger:
                    cell.fill = alter_row_fill
                else:
                    cell.fill = white_fill
                if i+idx < date_input_row:
                    cell.number_format = '0.0%'
                    cell.border = no_bottom_black_border
                    cell.font = percent_font
                    cell.alignment = align_bottom
                else:
                    cell.number_format = 'mmm d, yyyy'
                    cell.font = small_font
                    cell.alignment = align_top
                    cell.border = no_top_black_border

        i += rows_per_activity
        if alter_row_fill_trigger:
            alter_row_fill_trigger = False
        else:
            alter_row_fill_trigger = True

    # Add conditional formatting
    upper_left_cell = f'{get_col_letter_cached(_col_left_table + 2)}{_row_top_table + 3}'  # e.g., B1
    lower_right_cell = f'{get_column_letter(_col_left_table + n_progress_entry)}{_row_top_table + 2 + len(data) * rows_per_activity}'  # e.g., A1
    cell_range = f'{upper_left_cell}:{lower_right_cell}'
    current_cell = upper_left_cell
    previous_cell = f'{get_col_letter_cached(_col_left_table + 1)}{_row_top_table + 3}'

    # Formula: =B1<A1
    rule_formula = f'OR(AND({current_cell}<{previous_cell},{current_cell}<>""),AND({previous_cell}="",{current_cell}<>""))'

    # 6. Apply the Rule
    rule = FormulaRule(formula=[rule_formula], stopIfTrue=True, fill=red_fill)
    ws.conditional_formatting.add(cell_range, rule)

    # --- Insert Timeline Image ---
    try:
        img_path = resource_path('images/example_timeline.png')

        # 1. Open with PIL first to get exact pixel dimensions (ignores DPI metadata)
        with PILImage.open(img_path) as pil_img:
            actual_w, actual_h = pil_img.size

        # 2. Create the Excel Image
        img = Image(img_path)

        # 3. Enforce the dimensions to match pixels exactly (Fixes blurriness)
        img.width = actual_w
        img.height = actual_h

        img.anchor = f'J17'

        ws.add_image(img)

    except FileNotFoundError:
        print(f'Could not find image at {img_path}')
    except Exception as e:
        print(f'Error adding timeline image: {e}')

    return ws